#!/usr/bin/env python3
"""
Benchmark phenotype_matcher_v2 on NORMALIZATION only (no NER).

Uses the gold-standard text descriptions from the RAG-HPO benchmark as input,
and checks whether the tool maps each description to the correct HPO ID.

This isolates the normalization/concept-mapping step from entity recognition.

Evaluation modes:
  - exact:    predicted HPO ID == gold HPO ID
  - ancestor: predicted HPO ID is an ancestor or descendant of gold HPO ID
              (captures cases where a more general/specific term is predicted)
"""

import argparse
import json
import sys
import os
import time
from pathlib import Path
from collections import defaultdict

import openpyxl

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root / "tools" / "phenotype_matcher_v2" / "src"))

from phenotype_matcher_v2 import PhenotypeMatcher, MatcherConfig


def load_annotations(xlsx_path: str):
    """Load text descriptions and gold HPO IDs from the Excel file."""
    wb = openpyxl.load_workbook(xlsx_path)

    datasets = {}

    # CSC annotations: Patient ID, hpo_description, hpo_term
    ws = wb["CSC Manual Annotations"]
    csc = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[2]:
            desc = str(row[1]).strip()
            hpo = str(row[2]).strip()
            if hpo.startswith("HP:") and len(desc) > 2:
                csc.append({"description": desc, "gold_hpo": hpo, "case_id": str(row[0])})
    datasets["CSC"] = csc

    # GSC annotations: Patient ID, ID, hpo_description, hpo_term, Category
    ws = wb["GSC Manual Annotations "]
    gsc = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] and row[3]:
            desc = str(row[2]).strip()
            hpo = str(row[3]).strip()
            if hpo.startswith("HP:") and len(desc) > 2:
                gsc.append({"description": desc, "gold_hpo": hpo, "case_id": str(row[0])})
    datasets["GSC"] = gsc

    return datasets


def build_ancestor_map(hpo_path: str):
    """Build a mapping from each HPO term to all its ancestors using pronto."""
    import pronto
    ont = pronto.Ontology(hpo_path)
    ancestors = {}
    for term in ont.terms():
        tid = term.id
        anc = set()
        for parent in term.superclasses():
            if parent.id != tid:
                anc.add(parent.id)
        ancestors[tid] = anc
    return ancestors


def evaluate_normalization(annotations: list, matcher: PhenotypeMatcher,
                           ancestors: dict = None, dataset_name: str = ""):
    """
    For each gold (description, HPO ID) pair, run matcher on the description
    and check if the gold HPO ID appears in the output.
    """
    exact_match = 0
    ancestor_match = 0
    no_match = 0
    no_output = 0
    total = len(annotations)

    # Deduplicate: many descriptions repeat across cases
    unique_descs = {}
    for ann in annotations:
        desc = ann["description"].lower()
        if desc not in unique_descs:
            unique_descs[desc] = []
        unique_descs[desc].append(ann["gold_hpo"])

    print(f"\n{'='*60}")
    print(f"Normalization benchmark: {dataset_name}")
    print(f"Total annotations: {total}, unique descriptions: {len(unique_descs)}")
    print(f"{'='*60}", flush=True)

    # Cache matcher results for unique descriptions
    match_cache = {}
    start = time.time()
    for i, (desc, _) in enumerate(unique_descs.items()):
        try:
            output = matcher.match(desc)
            predicted = set(output.get_hpo_ids())
            # Also include excluded phenotypes for normalization eval
            predicted |= set(output.get_hpo_ids(excluded=True))
            match_cache[desc] = predicted
        except Exception as e:
            match_cache[desc] = set()

        if (i + 1) % 50 == 0 or (i + 1) == len(unique_descs):
            elapsed = time.time() - start
            rate = elapsed / (i + 1)
            remaining = rate * (len(unique_descs) - i - 1)
            print(f"  [{i+1:4d}/{len(unique_descs)}] "
                  f"{elapsed:.0f}s elapsed, ~{remaining:.0f}s left", flush=True)

    # Now evaluate all annotations
    mismatches = []
    for ann in annotations:
        desc = ann["description"].lower()
        gold = ann["gold_hpo"]
        predicted = match_cache.get(desc, set())

        if not predicted:
            no_output += 1
            mismatches.append({"desc": ann["description"], "gold": gold,
                               "predicted": [], "status": "no_output"})
        elif gold in predicted:
            exact_match += 1
        elif ancestors:
            # Check if any predicted term is an ancestor/descendant of gold
            gold_ancestors = ancestors.get(gold, set())
            found_ancestor = False
            for pred in predicted:
                pred_ancestors = ancestors.get(pred, set())
                if gold in pred_ancestors or pred in gold_ancestors:
                    found_ancestor = True
                    break
            if found_ancestor:
                ancestor_match += 1
            else:
                no_match += 1
                mismatches.append({"desc": ann["description"], "gold": gold,
                                   "predicted": list(predicted)[:5], "status": "wrong"})
        else:
            no_match += 1
            mismatches.append({"desc": ann["description"], "gold": gold,
                               "predicted": list(predicted)[:5], "status": "wrong"})

    elapsed = time.time() - start
    exact_acc = exact_match / total if total > 0 else 0
    ancestor_acc = (exact_match + ancestor_match) / total if total > 0 else 0
    coverage = (total - no_output) / total if total > 0 else 0

    print(f"\n--- {dataset_name} Normalization Results ---")
    print(f"Total annotations:    {total}")
    print(f"Exact match:          {exact_match} ({exact_acc:.1%})")
    if ancestors:
        print(f"Ancestor/desc match:  {ancestor_match} ({ancestor_match/total:.1%})")
        print(f"Combined accuracy:    {exact_match + ancestor_match} ({ancestor_acc:.1%})")
    print(f"Wrong:                {no_match} ({no_match/total:.1%})")
    print(f"No output:            {no_output} ({no_output/total:.1%})")
    print(f"Coverage:             {coverage:.1%}")
    print(f"Time:                 {elapsed:.1f}s ({elapsed/len(unique_descs):.2f}s/term)")

    # Show some mismatches for analysis
    print(f"\nSample mismatches (first 20):")
    for m in mismatches[:20]:
        print(f"  [{m['status']:>9s}] {m['desc']!r:40s} gold={m['gold']}  "
              f"pred={m['predicted'][:3]}")

    return {
        "dataset": dataset_name,
        "total": total,
        "unique_descriptions": len(unique_descs),
        "exact_match": exact_match,
        "exact_accuracy": round(exact_acc, 4),
        "ancestor_match": ancestor_match if ancestors else None,
        "combined_accuracy": round(ancestor_acc, 4) if ancestors else None,
        "wrong": no_match,
        "no_output": no_output,
        "coverage": round(coverage, 4),
        "time_seconds": round(elapsed, 1),
        "mismatches_sample": mismatches[:50],
    }


def main():
    parser = argparse.ArgumentParser(description="Benchmark normalization only")
    parser.add_argument("--no-llm", action="store_true",
                        help="Disable LLM validation")
    parser.add_argument("--no-ancestors", action="store_true",
                        help="Skip ancestor-aware evaluation")
    parser.add_argument("--dataset", choices=["csc", "gsc", "both"], default="both")
    args = parser.parse_args()

    base_dir = Path(__file__).parent
    xlsx_path = str(base_dir / "RAG-HPO_Tests_and_Data_Analysis.xlsx")

    api_key = None if args.no_llm else os.environ.get("OPENROUTER_API_KEY")
    mode = "offline" if args.no_llm or not api_key else "with LLM"
    print(f"Initializing phenotype_matcher_v2 [{mode}]...", flush=True)

    cfg = MatcherConfig(
        hpo_path=str(project_root / "ontology" / "hp.obo"),
        mondo_path=str(project_root / "ontology" / "mondo.obo"),
        hpoa_path=str(project_root / "data" / "phenotype.hpoa"),
        embedding_model="sapbert",
        device="cpu",
        api_key=api_key,
        llm_model="deepseek",
        debug=False,
        match_cache_size=10_000,
    )
    matcher = PhenotypeMatcher(cfg)
    print("Matcher initialized.", flush=True)

    # Build ancestor map for ancestor-aware evaluation
    ancestors = None
    if not args.no_ancestors:
        print("Building HPO ancestor map...", flush=True)
        ancestors = build_ancestor_map(cfg.hpo_path)
        print(f"Ancestor map: {len(ancestors)} terms", flush=True)

    # Load annotations
    datasets = load_annotations(xlsx_path)

    results = {}
    if args.dataset in ("csc", "both"):
        results["csc"] = evaluate_normalization(
            datasets["CSC"], matcher, ancestors, "CSC")
    if args.dataset in ("gsc", "both"):
        results["gsc"] = evaluate_normalization(
            datasets["GSC"], matcher, ancestors, "GSC")

    # Save results
    out_path = base_dir / "normalization_results.json"
    with open(out_path, "w") as f:
        json.dump(results, f, indent=2)
    print(f"\nResults saved to: {out_path}")


if __name__ == "__main__":
    main()
